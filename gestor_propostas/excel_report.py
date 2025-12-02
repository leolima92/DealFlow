import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .models import GestorPropostas


class ExcelReportGenerator:
    PASTA_RELATORIOS = "relatorios"

    @classmethod
    def _garantir_pasta(cls):
        if not os.path.isdir(cls.PASTA_RELATORIOS):
            os.makedirs(cls.PASTA_RELATORIOS, exist_ok=True)

    @classmethod
    def gerar_excel(cls, gestor: GestorPropostas, caminho_escolhido: Optional[str] = None) -> str:
        """
        Gera um Excel com duas abas:
        - Propostas: 1 linha por proposta (incluindo novos campos)
        - Itens: 1 linha por item, vinculado à proposta
        """
        cls._garantir_pasta()
        wb = Workbook()

        bold = Font(bold=True)
        center = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Aba 1: Propostas
        ws_prop = wb.active
        ws_prop.title = "Propostas"

        headers_prop = [
            "ID Proposta",
            "Título",
            "Cliente",
            "Documento",
            "Contato",
            "Status",
            "Data Criação",
            "Validade",
            "Responsável",
            "Condições Pagamento",
            "Subtotal",
            "Desconto",
            "Total",
        ]
        ws_prop.append(headers_prop)

        for col in range(1, len(headers_prop) + 1):
            cell = ws_prop.cell(row=1, column=col)
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border

        for p in gestor.listar_propostas():
            subtotal = p.calcular_subtotal()
            desconto = p.calcular_desconto()
            total = p.calcular_total()
            validade_str = p.validade.strftime("%Y-%m-%d") if p.validade else ""

            ws_prop.append([
                p.id,
                p.titulo,
                p.cliente.nome,
                p.cliente.documento,
                p.cliente.contato,
                p.status,
                p.data_criacao.strftime("%Y-%m-%d %H:%M:%S"),
                validade_str,
                p.responsavel,
                p.condicoes_pagamento,
                float(f"{subtotal:.2f}"),
                float(f"{desconto:.2f}"),
                float(f"{total:.2f}"),
            ])

        # Ajuste de largura
        for column_cells in ws_prop.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0
                         for cell in column_cells)
            ws_prop.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Aba 2: Itens
        ws_itens = wb.create_sheet("Itens")
        headers_itens = [
            "ID Proposta",
            "Título Proposta",
            "Cliente",
            "Descrição Item",
            "Quantidade",
            "Valor Unitário",
            "Total Item",
        ]
        ws_itens.append(headers_itens)
        for col in range(1, len(headers_itens) + 1):
            cell = ws_itens.cell(row=1, column=col)
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border

        for p in gestor.listar_propostas():
            for item in p.itens:
                ws_itens.append([
                    p.id,
                    p.titulo,
                    p.cliente.nome,
                    item.descricao,
                    item.quantidade,
                    float(f"{item.valor_unitario:.2f}"),
                    float(f"{item.total:.2f}"),
                ])

        for column_cells in ws_itens.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0
                         for cell in column_cells)
            ws_itens.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Caminho
        if caminho_escolhido:
            caminho = caminho_escolhido
        else:
            nome_arquivo = f"propostas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho = os.path.join(cls.PASTA_RELATORIOS, nome_arquivo)

        wb.save(caminho)
        return caminho
